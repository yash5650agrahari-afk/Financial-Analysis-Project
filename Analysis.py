# Financial Data Analysis Project
# Step 1 - Import Libraries

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
# Step 2 - Load the Dataset

df = pd.read_csv("Financials.csv")
print(df.head())

# Step 3 - Explore the Dataset
print(df.shape)
print(df.columns)
print(df.dtypes)
print(df.isnull().sum())

# Step 4 - Data Cleaning

# Fix columns names (remove extra spaces)
df.columns= df.columns.str.strip()

# check cleaned column names 
print(df.columns)

#Fix number columns (remove $ , symbol and convert to numbers)

cols_to_fix = ['Units Sold','Manufacturing Price','Sale Price','Gross Sales','Discounts','Sales','COGS','Profit' ]

for col in cols_to_fix:
    df[col] = df[col].str.replace('$','', regex=False)
    df[col] = df[col].str.replace('$','', regex=False)
    df[col] = df[col].str.replace(',','', regex=False)
    df[col] = pd.to_numeric(df[col],errors='coerce')

print(df.dtypes)

# Step 5 - Analysis 

## 1. Total Sales And Profit
total_sales = df['Sales'].sum()  # — selects only the Sales column (like clicking a column in Excel)
total_profit = df['Profit'].sum() #.sum() — adds up all values in that column (like SUM() in Excel)

print("Total Sales:", total_sales)
print("Total Profit:", total_profit)

## 2. Sales by Country

sales_by_country = df.groupby('Country')['Sales'].sum().sort_values(ascending=False)
print(sales_by_country)

## 3. Profit by Product
profit_by_product = df.groupby('Product')['Profit'].sum().sort_values(ascending=False)
print(profit_by_product)

## 4. Profit by Segment

profit_by_segment = df.groupby('Segment')['Profit'].sum().sort_values(ascending=False)
print(profit_by_segment)


## 5. Sales by Year
sales_by_year = df.groupby('Year')['Sales'].sum().sort_values(ascending=False)
print(sales_by_year)

# Step 6 - Visualization

# 1.Sales by Country - Bar Chart
plt.figure(figsize=(10,6))
sns.barplot(x=sales_by_country.index, y=sales_by_country.values)
plt.title('Total Sales by Country')
plt.xlabel('Country')
plt.ylabel('Sales')
plt.tight_layout()
plt.savefig('sales_by_country.png')
plt.show()
print("Chart 1 saved!")


# # 2. Profit  by Product - Bar Chart

plt.figure(figsize=(10,6))
sns.barplot(x=profit_by_product.index, y=profit_by_product.values)
plt.title('Total Profit by Product')
plt.xlabel('Product')
plt.ylabel('Profit')
plt.tight_layout()
plt.savefig('profit_by_product.png')
plt.show()
print("Chart 2 saved!")

# # 3. Profit by Segment - Bar Chart
plt.figure(figsize=(10, 6))
sns.barplot(x=profit_by_segment.index, y=profit_by_segment.values)
plt.title('Total Profit by Segment')
plt.xlabel('Segment')
plt.ylabel('Profit')
plt.tight_layout()
plt.savefig('profit_by_segment.png')
plt.show()
print("Chart 3 saved!")


# 4. Profit by Year - Bar Chart
plt.figure(figsize=(10, 6))
sns.barplot(x=sales_by_year.index, y=sales_by_year.values)
plt.title('Total Sales by Year')
plt.xlabel('Year')
plt.ylabel('Sales')
plt.tight_layout()
plt.savefig('sales_by_year.png')
plt.show()
print("Chart 4 saved!")

# 5. Profit Margin by Country - Bar Chart
df['Profit Margin'] = (df['Profit'] / df['Sales']) * 100

profit_margin = df.groupby('Country')['Profit Margin'].mean().sort_values(ascending=False)

plt.figure(figsize=(10, 6))
sns.barplot(x=profit_margin.index, y=profit_margin.values)
plt.title('Average Profit Margin by Country')
plt.xlabel('Country')
plt.ylabel('Profit Margin %')
plt.tight_layout()
plt.savefig('profit_margin_by_country.png')
plt.show()
print("Chart 5 saved!")

# Step 7 - Export to Excel

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Financial Report"

# Heading
ws['A1'] = "Financial Data Analysis Report"
ws['A2'] = "Summary of Key Insights"

# Total Sales and Profit
ws['A4'] = "Metric"
ws['B4'] = "Value"
ws['A5'] = "Total Sales"
ws['B5'] = round(total_sales, 2)
ws['A6'] = "Total Profit"
ws['B6'] = round(total_profit, 2)

# Sales by Country
ws['A9'] = "Country"
ws['B9'] = "Sales"
for i, (country, sales) in enumerate(sales_by_country.items(), start=10):
    ws[f'A{i}'] = country
    ws[f'B{i}'] = round(sales, 2)

# Profit by Product
ws['D9'] = "Product"
ws['E9'] = "Profit"
for i, (product, profit) in enumerate(profit_by_product.items(), start=10):
    ws[f'D{i}'] = product
    ws[f'E{i}'] = round(profit, 2)

# Profit by Segment
ws['G9'] = "Segment"
ws['H9'] = "Profit"
for i, (segment, profit) in enumerate(profit_by_segment.items(), start=10):
    ws[f'G{i}'] = segment
    ws[f'H{i}'] = round(profit, 2)

wb.save("Financial_Report.xlsx")
print("Excel Report saved!")